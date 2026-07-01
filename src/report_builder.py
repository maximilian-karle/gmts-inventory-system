"""
report_builder.py
==================
Erstellt den finalen Excel-Report fuer einen Analyse-Lauf.

Aufbau: EIN kombinierter Reiter pro Datenquelle/Phase (Entscheidung Max,
24.06.2026) statt mehrerer Detail-Reiter. Fuer den ZMLAG-Lauf heisst der
Reiter "ZMLAG_Bestand" und enthaelt alle Kennzahlen aus diesem Run
(Materialstammdaten, Stichtags-Bestandswerte, Bestandsentwicklung inkl.
tagesaktuellem "Laufender Wert", ABC/XYZ-Risikoeinstufung).

Hintergrund: Mit jeder weiteren Projektphase (Reichweite, Trend/MVER,
Wiederbeschaffungszeit, Fruehwarnsystem) kommt ein zusaetzlicher, eigener
Reiter in derselben Arbeitsmappe hinzu (z.B. "Reichweite", "Trend_MVER",
"Wiederbeschaffung", "Fruehwarnsystem") - bewusst schmal gehalten statt
mit mehreren Detail-Reitern je Phase, damit die Datei bei wachsendem
Funktionsumfang uebersichtlich bleibt.

Bewusste Trennung von Berechnung (stock_analysis.py) und Aufbereitung/
Formatierung (dieses Modul), damit Aenderungen am Excel-Layout nie die
fachliche Berechnungslogik beruehren muessen.

Zusatzfunktion build_consolidated_report() (siehe main_all.py): erzeugt
einen zusaetzlichen, technologieuebergreifenden Report mit allen
Technologien in einer gemeinsamen Tabelle - unabhaengig von den weiterhin
bestehenden Einzelreports je Technologie.
"""

from __future__ import annotations

from dataclasses import dataclass, field

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList

from config import RunConfig
from data_loader import get_period_columns

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Deutsches Zahlenformat fuer Waehrungsspalten (Tausenderpunkt, Komma als
# Dezimaltrennzeichen, Euro-Zeichen) bzw. Prozentspalten.
CURRENCY_FORMAT = "#,##0.00 €"
PERCENT_FORMAT = "0.00%"

# Name des kombinierten Reiters fuer den ZMLAG-Lauf. Kuenftige Phasen
# (Reichweite, Trend/MVER, Wiederbeschaffung, Fruehwarnsystem) bekommen
# jeweils einen eigenen, analog benannten Reiter in derselben Arbeitsmappe.
ZMLAG_SHEET_NAME = "ZMLAG_Bestand"

# Materialstammdaten + Stichtags-Werte (letztere werden in build_report
# dynamisch ergaenzt, da ihre Spaltennamen pro Export variieren) + alle
# Kennzahlen aus stock_analysis.py, kombiniert in einem Reiter.
ZMLAG_BASE_COLUMNS = [
    "Material",
    "Materialkurztext",
    "Normbezeichnung",
    "Disponent",
    "Technologie",
]

ZMLAG_CALCULATED_COLUMNS = [
    "Laufender Wert",
    "Bestandswert_Start",
    "Bestandswert_Letzter_Stichtag",
    "Bestandswert_Ende",
    "Veraenderung_Absolut",
    "Veraenderung_Prozent",
    "ABC-Kennzeichen",
    "ABCXYZ-Kennzeichen",
    "Risiko_Einstufung",
]

# Spalten, die als Waehrung formatiert werden sollen (Stichtagsspalten werden
# in build_report dynamisch hinzugefuegt, da ihre Namen pro Export variieren).
CURRENCY_COLUMNS_FIXED = [
    "Laufender Wert",
    "Bestandswert_Start",
    "Bestandswert_Letzter_Stichtag",
    "Bestandswert_Ende",
    "Veraenderung_Absolut",
]

PERCENT_COLUMNS = [
    "Veraenderung_Prozent",
]


# Name des Reiters im konsolidierten Multi-Technologie-Report, der alle
# Technologien in einer gemeinsamen Tabelle zusammenfasst (siehe
# build_consolidated_report). Je-Technologie-Einzelreports (build_report)
# bleiben davon unberuehrt und werden weiterhin separat erzeugt.
CONSOLIDATED_SHEET_NAME = "Alle_Technologien"


# ---------------------------------------------------------------------------
# Reichweite-Reiter (Phase 2, ab 24.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter in derselben Arbeitsmappe wie ZMLAG_Bestand (siehe Modul-
# Docstring: "ein kombinierter Reiter pro Phase"). Enthaelt die Ergebnisse
# aus coverage_analysis.build_coverage_table() - Bestandsmenge (Stueck) aus
# SE16XXL, Ø-Verbrauch (12 Monate) aus MVER, Planungsfeld-Override und die
# daraus berechnete Reichweite (Stock-only).
COVERAGE_SHEET_NAME = "Reichweite"

COVERAGE_COLUMNS = [
    "Material",
    "MatStatus",
    "MatStatus_Bezeichnung",
    "MatStatus_Kritisch",
    "Bestandsmenge",
    "Ø_Verbrauch_Fenster",
    "Anzahl_Monate_Verfuegbar",
    "Planung_Verbrauch_Override",
    "Verbrauch_Fuer_Reichweite",
    "Reichweite_Monate_StockOnly",
]

# Spalten im Reichweite-Reiter, die ein Zahlenformat mit 2 Nachkommastellen
# erhalten (keine Waehrung, da Stueck/Monate - daher eigenes, einfaches Format
# statt CURRENCY_FORMAT).
COVERAGE_DECIMAL_FORMAT = "#,##0.00"
COVERAGE_DECIMAL_COLUMNS = [
    "Bestandsmenge",
    "Ø_Verbrauch_Fenster",
    "Planung_Verbrauch_Override",
    "Verbrauch_Fuer_Reichweite",
    "Reichweite_Monate_StockOnly",
]


# ---------------------------------------------------------------------------
# Trend_MVER-Reiter (Phase 3, ab 25.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter, analog zum Reichweite-Reiter aus Phase 2. Enthaelt sowohl
# die ROHE monatliche Verbrauchs-Zeitreihe (eine Spalte je Monat, Format
# 'YYYY-MM') als auch die daraus abgeleiteten Kennzahlen aus
# trend_analysis.build_trend_table() - erfuellt die explizite Anforderung
# aus "Initiale Informationen" ("36 Monate Historie monatlich in einem
# getrennten Reiter, damit Trends, Sondereffekte und echte Entwicklung
# erkannt werden").
TREND_SHEET_NAME = "Trend_MVER"

# Kennzahlen-Spalten aus trend_analysis.build_trend_table(), die NACH den
# (dynamischen, je Lauf wechselnden) Monatsspalten erscheinen.
TREND_METRIC_COLUMNS = [
    "Trend_Steigung_Stueck_Monat",
    "Trend_Einstufung",
    "Sondereffekt_Monate",
    "Anzahl_Sondereffekte",
    "Verbrauch_Streuung_MAD",
    "Saisonalitaet_Hinweis",
    "Anzahl_Monate_Verfuegbar",
]

# Spalten im Trend_MVER-Reiter, die als einfache Dezimalzahl formatiert
# werden (Stueck bzw. Stueck/Monat - kein Waehrungsformat). Die rohen
# Monatsspalten werden separat behandelt (siehe build_report/_apply_formatting),
# da ihre Namen dynamisch sind und nicht vorab feststehen.
TREND_DECIMAL_COLUMNS = [
    "Trend_Steigung_Stueck_Monat",
    "Verbrauch_Streuung_MAD",
]

# Fuellfarbe fuer als Sondereffekt erkannte Zellen in den Monatsspalten
# (helles Orange) - macht den herausgerechneten Peak optisch erkennbar,
# OHNE den Zahlenwert selbst zu veraendern (Klaerung mit Max, 24.06.2026).
OUTLIER_CELL_FILL = PatternFill(start_color="FFD9A6", end_color="FFD9A6", fill_type="solid")


# ---------------------------------------------------------------------------
# Wiederbeschaffung-Reiter (Phase 4, ab 25.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter, analog zum Reichweite-Reiter aus Phase 2 (Entscheidung
# Max, 25.06.2026: eigener Reiter statt Integration in 'Reichweite').
# Enthaelt das Ergebnis von dispo_abcxyz_loader.build_lead_time_table() -
# Wiederbeschaffungszeit (Planlieferzeit + WE-Bearbeitungszeit, Whitepaper
# Kap. 3.2) sowie die uebrigen, bereits jetzt vollstaendig eingelesenen
# Dispo_ABCXYZ-Felder (Vari.Koeff, Sicherheitsbestand, Meldebestand, Preise,
# ABCXYZ-Kennzeichen etc.), die erst in Phase 5/6 fachlich genutzt werden,
# aber bereits hier sichtbar gemacht werden, statt den Reiter spaeter
# erneut anzufassen.
LEAD_TIME_SHEET_NAME = "Wiederbeschaffung"

LEAD_TIME_COLUMNS = [
    "Material",
    "ABCXYZ_Kombiniert",
    "MatStatus",
    "Disponent",
    "Einkaeufergruppe",
    "Beschaffungsart",
    "Planlieferzeit_Tage",
    "WE_Bearbeitungszeit_Tage",
    "Wiederbeschaffungszeit_Tage",
    "Variationskoeffizient",
    "Sicherheitsbestand_Soll",
    "Mindest_Sicherheitsbestand",
    "Meldebestand",
    "Standardpreis",
    "Gleitender_Durchschnittspreis",
    "Mindest_Losgroesse",
    "Hoechst_Losgroesse",
    "Feste_Losgroesse",
    "Rundungswert",
    "Hoechstbestand",
]

# Spalten im Wiederbeschaffung-Reiter, die als einfache Dezimalzahl (2
# Nachkommastellen, kein Waehrungsformat) formatiert werden - analog zu
# COVERAGE_DECIMAL_COLUMNS. Preise (Standardpreis, Gleitender_Durchschnitts
# preis) erhalten bewusst CURRENCY_FORMAT statt dieses Formats (siehe
# LEAD_TIME_CURRENCY_COLUMNS unten), da es sich um echte Euro-Betraege
# handelt, nicht um Tage/Stueck/Verhaeltniszahlen.
LEAD_TIME_DECIMAL_COLUMNS = [
    "Planlieferzeit_Tage",
    "WE_Bearbeitungszeit_Tage",
    "Wiederbeschaffungszeit_Tage",
    "Variationskoeffizient",
    "Sicherheitsbestand_Soll",
    "Mindest_Sicherheitsbestand",
    "Meldebestand",
    "Mindest_Losgroesse",
    "Hoechst_Losgroesse",
    "Feste_Losgroesse",
    "Rundungswert",
    "Hoechstbestand",
]

LEAD_TIME_CURRENCY_COLUMNS = [
    "Standardpreis",
    "Gleitender_Durchschnittspreis",
]


# ---------------------------------------------------------------------------
# Executive_Summary-Reiter (Phase 6, ab 29.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter, analog zum bisherigen Aufbau je Phase. Enthaelt das
# Ergebnis von executive_summary.build_executive_summary() - eine Zeile pro
# Material, verdichtet aus ZMLAG/Reichweite/Trend/Fruehwarnsystem/
# Simulation zu den wichtigsten Kennzahlen (Klaerung mit Max, 29.06.2026 -
# "roter Faden" fuer wachsenden Reiter-Umfang, siehe Projektstatus.md
# Abschnitt 4f). KEIN Ersatz fuer die Detail-Reiter, sondern ein
# vorangestellter Ueberblick - alle Detail-Reiter bleiben unveraendert
# bestehen.
EXECUTIVE_SUMMARY_SHEET_NAME = "Executive_Summary"

EXECUTIVE_SUMMARY_COLUMNS = [
    "Material",
    "Materialkurztext",
    "ABC-Kennzeichen",
    "ABCXYZ-Kennzeichen",
    "Bestandswert_EUR",
    "Bestandsmenge_Stueck",
    "Reichweite_Monate",
    "MatStatus_Bezeichnung",
    "Trend_Einstufung",
    "Meldebestand_ROP",
    "Delta_Meldebestand",
    "Fehlmengen_Wahrscheinlichkeit_Horizont",
    "Working_Capital_Reduktion",
    "ROI_Prozent",
    "Prioritaet",
]

EXECUTIVE_SUMMARY_DECIMAL_COLUMNS = [
    "Bestandsmenge_Stueck",
    "Reichweite_Monate",
    "Meldebestand_ROP",
    "Delta_Meldebestand",
]

EXECUTIVE_SUMMARY_CURRENCY_COLUMNS = [
    "Bestandswert_EUR",
    "Working_Capital_Reduktion",
]

EXECUTIVE_SUMMARY_PERCENT_COLUMNS = [
    "Fehlmengen_Wahrscheinlichkeit_Horizont",
    "ROI_Prozent",
]

# Hinweis: die zugehoerige EXECUTIVE_SUMMARY_SHEET_SPEC-Instanz steht weiter
# unten, gemeinsam mit den uebrigen SheetSpec-Instanzen (SheetSpec ist an
# dieser Stelle im Modul noch nicht definiert).

# Farbliche Markierung der Spalte 'Prioritaet' (Ampel-Logik), angewendet in
# _apply_priority_highlighting() - bewusst eigene, kleine Funktion statt
# Erweiterung von _apply_formatting() um einen Spezialfall, der nur diesen
# einen Reiter betrifft (analog zur Begruendung fuer eigene Trend-Sheet-
# Funktionen oben: ein generischer Mechanismus fuer einen Einzelfall wuerde
# selbst wieder unuebersichtlich).
PRIORITY_FILL_COLORS = {
    "Kritisch": PatternFill(start_color="F4A6A6", end_color="F4A6A6", fill_type="solid"),
    "Erhoeht": PatternFill(start_color="FCE5A6", end_color="FCE5A6", fill_type="solid"),
    "Unauffaellig": PatternFill(start_color="C6E5B3", end_color="C6E5B3", fill_type="solid"),
    "Unbekannt": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}


# ---------------------------------------------------------------------------
# Dashboard-Reiter (Phase 6, ab 29.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter NUR mit Diagrammen (keine eigene Datentabelle, abgesehen
# von kleinen Hilfstabellen, die als Datenquelle fuer die Charts dienen -
# siehe _build_dashboard_sheet()). Bewusst getrennt von Executive_Summary
# (Entscheidung Max, 29.06.2026: eine Tabelle + ein separater Dashboard-
# Reiter, nicht beides vermischt) - die Tabellenreiter bleiben dadurch
# unveraendert maschinenlesbar (z.B. als Power-BI-Quelle, siehe
# Projektstatus.md Abschnitt 4f), waehrend der Dashboard-Reiter frei
# layoutet werden kann, ohne die Datenstruktur zu stoeren.
#
# openpyxl liest die Chart-Datenquelle immer aus Zellen IM SELBEN Workbook -
# die Hilfstabellen (Aggregationen je Technologie, Top-N-Ranglisten etc.)
# werden daher auf dem Dashboard-Reiter selbst geschrieben, in einem
# Bereich unterhalb/rechts der Charts (siehe DASHBOARD_DATA_START_ROW),
# damit sie im PDF-/Bildschirm-Layout nicht stoeren, aber bei Bedarf
# nachvollziehbar bleiben (kein verstecktes Hilfsblatt, das beim
# Filtern/Kopieren verloren gehen koennte).
DASHBOARD_SHEET_NAME = "Dashboard"

# Erste Datenzeile fuer die Hilfstabellen auf dem Dashboard-Reiter (Zeile 1
# bleibt fuer Titel/Kontext frei, Diagramme werden ueber den Hilfstabellen
# platziert - siehe _build_dashboard_sheet()).
DASHBOARD_HELPER_TABLE_START_ROW = 2




# ---------------------------------------------------------------------------
# Simulation_Verbrauch-Reiter (Baustein A, ab 25.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter fuer die Monte-Carlo-Verbrauchsprognose-Simulation (siehe
# simulation_analysis.py). Bewusst eigener Reiter statt Integration in
# 'Reichweite' oder 'Trend_MVER' (Entscheidung Claude, 25.06.2026, analog zur
# bisherigen Praxis je neuem Baustein einen eigenen Reiter anzulegen, siehe
# z.B. 'Wiederbeschaffung' fuer Phase 4) - die Simulation konsumiert BEIDE
# Quellen (Bestandsmenge aus Reichweite, Trend/Streuung aus Trend_MVER), ist
# aber inhaltlich ein eigenstaendiges Ergebnis (Wahrscheinlichkeitsverteilung
# statt einzelner Kennzahl).
SIMULATION_SHEET_NAME = "Simulation_Verbrauch"

SIMULATION_COLUMNS = [
    "Material",
    "Bestandsmenge_Start",
    "Verbrauch_Mittelwert_Basis",
    "Trend_Steigung_Stueck_Monat",
    "Verbrauch_Streuung_MAD",
    "Reichweite_Monate_P10",
    "Reichweite_Monate_P50",
    "Reichweite_Monate_P90",
    "Fehlmengen_Wahrscheinlichkeit_Horizont",
    "Erwarteter_Bestand_Horizont_Ende",
    "Planungshorizont_Monate",
    "Anzahl_Simulationen",
]

# Spalten im Simulation_Verbrauch-Reiter mit einfachem Dezimalformat (Stueck/
# Monate, kein Waehrungsformat) - analog zu COVERAGE_DECIMAL_COLUMNS.
SIMULATION_DECIMAL_COLUMNS = [
    "Bestandsmenge_Start",
    "Verbrauch_Mittelwert_Basis",
    "Trend_Steigung_Stueck_Monat",
    "Verbrauch_Streuung_MAD",
    "Reichweite_Monate_P10",
    "Reichweite_Monate_P50",
    "Reichweite_Monate_P90",
    "Erwarteter_Bestand_Horizont_Ende",
]

# Fehlmengen_Wahrscheinlichkeit_Horizont ist ein Anteil (0.0-1.0) und erhaelt
# daher PERCENT_FORMAT statt eines Dezimalformats - separat aufgefuehrt, da
# PERCENT_COLUMNS (siehe oben) bereits fix mit den ABC/XYZ-Risikospalten aus
# Phase 1 belegt ist und hier nicht global erweitert werden soll.
SIMULATION_PERCENT_COLUMNS = [
    "Fehlmengen_Wahrscheinlichkeit_Horizont",
]


# ---------------------------------------------------------------------------
# Fruehwarnsystem-Reiter (Phase 5, ab 29.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter fuer Safety Stock und Meldebestand (ROP), siehe
# safety_stock.py. Spaltenreihenfolge bewusst als Herleitungskette von
# links nach rechts (Klaerung mit Max, 29.06.2026 - "roter Faden"): Klasse
# -> Ziel-Servicegrad -> z-Wert -> Streuung -> Wiederbeschaffungszeit ->
# Verbrauch -> Safety_Stock -> Meldebestand_ROP -> SAP-Vergleich -> Delta.
# Die Klartext-Spalte 'Berechnungsbasis' fasst diese Kette je Material in
# einem Satz zusammen, fuer die COO-Kommunikation (analog zum Zweck von
# Trend_Details in Phase 3 - Nachvollziehbarkeit ohne Rueckrechnung von
# Hand). SAP-Vergleichsspalten (SAP_Sicherheitsbestand_Soll, SAP_Meldebestand,
# Delta_*) sind in SAFETY_STOCK_COLUMNS gelistet, aber nur vorhanden, wenn
# safety_stock.build_safety_stock_table() sie tatsaechlich erzeugt hat
# (abhaengig davon, ob lead_time_df die SAP-Felder enthielt) - der
# bestehende "nur verfuegbare Spalten schreiben"-Mechanismus in
# _build_single_sheet()/_build_consolidated_sheet() greift hier identisch
# wie bei allen anderen Reitern.
SAFETY_STOCK_SHEET_NAME = "Fruehwarnsystem"

SAFETY_STOCK_COLUMNS = [
    "Material",
    "ABCXYZ_Kombiniert",
    "Ziel_Servicegrad",
    "Z_Wert",
    "Verbrauch_Streuung_Sigma",
    "Wiederbeschaffungszeit_Monate",
    "Verbrauch_Fuer_Reichweite",
    "Safety_Stock",
    "Meldebestand_ROP",
    "Berechnungsbasis",
    "SAP_Sicherheitsbestand_Soll",
    "Delta_Safety_Stock",
    "SAP_Meldebestand",
    "Delta_Meldebestand",
]

# Spalten im Fruehwarnsystem-Reiter mit einfachem Dezimalformat (Stueck/
# Monate, kein Waehrungsformat) - analog zu COVERAGE_DECIMAL_COLUMNS.
# 'Berechnungsbasis' ist Text und bekommt KEIN Zahlenformat.
SAFETY_STOCK_DECIMAL_COLUMNS = [
    "Z_Wert",
    "Verbrauch_Streuung_Sigma",
    "Wiederbeschaffungszeit_Monate",
    "Verbrauch_Fuer_Reichweite",
    "Safety_Stock",
    "Meldebestand_ROP",
    "SAP_Sicherheitsbestand_Soll",
    "Delta_Safety_Stock",
    "SAP_Meldebestand",
    "Delta_Meldebestand",
]

# Ziel_Servicegrad ist ein Anteil (0.90-0.99) und erhaelt PERCENT_FORMAT -
# separat aufgefuehrt aus demselben Grund wie SIMULATION_PERCENT_COLUMNS.
SAFETY_STOCK_PERCENT_COLUMNS = [
    "Ziel_Servicegrad",
]


# ---------------------------------------------------------------------------
# Working_Capital-Reiter (Phase 6 Fortsetzung, ab 30.06.2026)
# ---------------------------------------------------------------------------
# Eigener Reiter fuer die Working-Capital-/ROI-Berechnung, siehe
# working_capital.py und Projektstatus.md Abschnitt 4i fuer die
# Designklaerung (insb. die Vereinfachung "Optimal = Meldebestand_ROP" statt
# einer vollwertigen Policy-Optimierungssimulation, und die Scope-
# Eingrenzung v1 ohne Shortage_Cost/Ordering_Cost). Spaltenreihenfolge
# bewusst als Herleitungskette von links nach rechts (analog zum
# Fruehwarnsystem-Reiter oben): Bestandsmenge/ROP/Preis -> Working Capital
# Aktuell/Optimal/Reduktion -> Holding Cost -> Annual Savings -> ROI.
WORKING_CAPITAL_SHEET_NAME = "Working_Capital"

WORKING_CAPITAL_COLUMNS = [
    "Material",
    "Bestandsmenge_Stueck",
    "Meldebestand_ROP_Stueck",
    "Standardpreis",
    "Working_Capital_Aktuell",
    "Working_Capital_Optimal",
    "Working_Capital_Reduktion",
    "Holding_Cost_Aktuell",
    "Holding_Cost_Optimal",
    "Annual_Savings",
    "ROI_Prozent",
]

WORKING_CAPITAL_DECIMAL_COLUMNS = [
    "Bestandsmenge_Stueck",
    "Meldebestand_ROP_Stueck",
]

WORKING_CAPITAL_CURRENCY_COLUMNS = [
    "Standardpreis",
    "Working_Capital_Aktuell",
    "Working_Capital_Optimal",
    "Working_Capital_Reduktion",
    "Holding_Cost_Aktuell",
    "Holding_Cost_Optimal",
    "Annual_Savings",
]

WORKING_CAPITAL_PERCENT_COLUMNS = [
    "ROI_Prozent",
]


# ---------------------------------------------------------------------------
# Forecast_MVER-Reiter (Modellauswahl-Forecast, ab 29.06.2026)
# ---------------------------------------------------------------------------
# Eigener, paralleler Reiter neben Trend_MVER (Klaerung mit Max, 29.06.2026):
# Max hat ein Referenz-Snippet aus seinem Python-in-Excel-Inventory-Tool
# gezeigt (Naive/SES/Holt/Holt-Winters-Modellauswahl per RMSE auf einem
# Train/Test-Split, siehe forecast_analysis.py) und moechte diese Erweiterung
# UMSETZEN, NICHT als Ersatz fuer die bestehende lineare Trend-Steigung aus
# Phase 3 (trend_analysis.py bleibt unveraendert), sondern als zusaetzlicher,
# alternativer Blick auf denselben MVER-Verbrauch - mit explizitem
# 12-Monats-Punktforecast und Modell-Transparenz (welches Modell, wie gut
# laut RMSE/Bias, welche Glaettungsparameter).
#
# Bewusst LANGFORMAT (eine Zeile pro Material+Forecast-Monat, identisch zur
# Struktur des Excel-Vorbilds und zu simulation_analysis.py's Aufbau), NICHT
# pivotiert wie Trend_MVER - hier gibt es keine rohen Vergangenheitsmonate
# zum Anzeigen, nur den Forecast selbst, daher passt die kompaktere
# Langform besser und der Reiter passt unveraendert in den generischen
# SheetSpec-Mechanismus (keine dynamischen Spalten, kein Cell-Highlighting -
# anders als Trend_MVER/Trend_Details).
FORECAST_SHEET_NAME = "Forecast_MVER"

FORECAST_COLUMNS = [
    "Material",
    "Forecast_Monat",
    "Forecast_Verbrauch",
    "Best_Model",
    "RMSE",
    "Bias",
    "Alpha",
    "Beta",
    "Gamma",
    "Total_Data_Points",
    "Train_Months",
    "Test_Months",
]

# Spalten im Forecast_MVER-Reiter mit einfachem Dezimalformat (Stueck bzw.
# dimensionslose Modellparameter, kein Waehrungsformat) - analog zu
# SIMULATION_DECIMAL_COLUMNS. 'Best_Model' ist Text und bekommt KEIN
# Zahlenformat, 'Total_Data_Points'/'Train_Months'/'Test_Months' sind
# ganzzahlige Metadaten-Spalten (analog zu Planungshorizont_Monate/
# Anzahl_Simulationen in simulation_analysis.py) - bewusst OHNE
# Dezimalformat, da Nachkommastellen bei einer Monatsanzahl irritieren
# wuerden.
FORECAST_DECIMAL_COLUMNS = [
    "Forecast_Verbrauch",
    "RMSE",
    "Bias",
    "Alpha",
    "Beta",
    "Gamma",
]


# ---------------------------------------------------------------------------
# Trend_Details-Reiter (Transparenz-Erweiterung, ab 25.06.2026)
# ---------------------------------------------------------------------------
# Eigener, zusaetzlicher Reiter neben Trend_MVER (Klaerung mit Max,
# 25.06.2026): macht die Sondereffekt-Erkennung je Material UND je Monat
# vollstaendig nachvollziehbar - Median, MAD (gesamt) sowie der modifizierte
# Z-Score JEDES einzelnen Monats, nicht nur die bereits abgeleiteten
# Endergebnisse (Sondereffekt_Monate, Verbrauch_Streuung_MAD) aus
# Trend_MVER. Hintergrund: Max moechte Analyseergebnisse gegenueber dem COO
# fundiert erklaeren koennen ("wie genau kommt dieser Wert zustande"), nicht
# nur das Ergebnis selbst zeigen. Bewusst ein EIGENER Reiter (nicht in
# Trend_MVER integriert), da sich die Spaltenzahl durch die zusaetzliche
# Z-Score-Spalte je Monat verdoppelt - das waere im ohnehin breiten
# Trend_MVER-Reiter unuebersichtlich.
TREND_DETAILS_SHEET_NAME = "Trend_Details"

# Kennzahlen-Spalten aus trend_analysis.build_trend_table() (details_df),
# die VOR den dynamischen Monatsspalten erscheinen.
TREND_DETAILS_METRIC_COLUMNS = [
    "Verbrauch_Median",
    "Verbrauch_Streuung_MAD",
]


# ---------------------------------------------------------------------------
# Generischer Section-Mechanismus (Refactoring, 29.06.2026)
# ---------------------------------------------------------------------------
# Hintergrund: Mit jeder Phase (Reichweite, Wiederbeschaffung, Simulation,
# kuenftig Fruehwarnsystem) kam in build_report()/build_consolidated_report()
# ein weiterer optionaler Parameter sowie ein nahezu identischer Codeblock
# hinzu (Spalten filtern -> Reiter schreiben -> Format-Listen ergaenzen ->
# im konsolidierten Fall zusaetzlich 'Technologie'-Spalte einfuegen). Diese
# Wiederholung wird hier durch EINE deklarative Beschreibung pro Reiter
# (SheetSpec) plus zwei generische Funktionen ersetzt.
#
# Bewusst NICHT in diesen Mechanismus gepresst: TREND_SHEET_NAME und
# TREND_DETAILS_SHEET_NAME (Entscheidung Claude, 29.06.2026, in Abstimmung
# mit Max). Beide unterscheiden sich strukturell von den uebrigen Reitern
# (dynamische Monatsspalten statt fester Spaltenliste, Z-Score-Paarspalten,
# zellweises Outlier-Highlighting via Material x Spalte) - eine generische
# Struktur, die auch diese Faelle abdeckt, wuerde Spezialfelder enthalten,
# die nur diese zwei von sechs Reitern nutzen, und waere damit selbst
# wieder ein kleiner Monolith. Die beiden Trend-Reiter bleiben daher als
# eigene, klar markierte Funktionen bestehen (siehe _build_trend_sheet(),
# _build_trend_details_sheet() weiter unten) - identisch zur bisherigen
# Logik, nur aus build_report()/build_consolidated_report() ausgelagert.
@dataclass(frozen=True)
class SheetSpec:
    """Deklarative Beschreibung eines 'einfachen' Reiters (statische
    Spaltenliste, kein dynamisches Spaltenlayout, kein Cell-Highlighting).

    Eine SheetSpec pro Phase/Reiter (Reichweite, Wiederbeschaffung,
    Simulation, Fruehwarnsystem, ...) ersetzt den bisher je Reiter
    wiederholten Codeblock in build_report()/build_consolidated_report().
    """

    sheet_name: str
    columns: list[str]
    decimal_columns: list[str] = field(default_factory=list)
    currency_columns: list[str] = field(default_factory=list)
    percent_columns: list[str] = field(default_factory=list)


@dataclass
class _FormatCollector:
    """Sammelt waehrend des Schreibens aller Reiter die Spaltennamen, die am
    Ende gemeinsam an _apply_formatting() uebergeben werden. Ersetzt die
    bisher in build_report()/build_consolidated_report() von Hand mit '+'
    zusammengefuehrten einzelnen Listen (currency_columns, decimal_columns,
    percent_columns) durch EIN Objekt, das jede SheetSpec-Verarbeitung
    direkt ergaenzen kann, ohne dass die Aufrufstelle jede neue Phase
    erneut per Hand verdrahten muss.
    """

    currency_columns: list[str] = field(default_factory=list)
    decimal_columns: list[str] = field(default_factory=list)
    percent_columns: list[str] = field(default_factory=list)


def _build_single_sheet(
    spec: SheetSpec,
    source_df: pd.DataFrame | None,
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> None:
    """Schreibt EINEN Reiter fuer den Einzel-Report (build_report), sofern
    source_df gesetzt ist, und registriert die zugehoerigen Format-Spalten
    in fmt. Tut nichts, falls source_df None ist (Phase fuer diesen Lauf
    nicht aktiv) - haelt build_report() abwaertskompatibel.
    """
    if source_df is None:
        return

    available_columns = [col for col in spec.columns if col in source_df.columns]
    _write_sheet(source_df[available_columns], writer, spec.sheet_name)

    fmt.decimal_columns += [col for col in spec.decimal_columns if col in available_columns]
    fmt.currency_columns += [col for col in spec.currency_columns if col in available_columns]
    fmt.percent_columns += [col for col in spec.percent_columns if col in available_columns]


def _build_consolidated_sheet(
    spec: SheetSpec,
    per_technology_source_dfs: dict[str, pd.DataFrame] | None,
    per_technology_dfs: dict[str, pd.DataFrame],
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> None:
    """Schreibt EINEN konsolidierten Reiter (build_consolidated_report) ueber
    alle Technologien hinweg, sofern per_technology_source_dfs gesetzt und
    nicht leer ist. Ergaenzt je Technologie die Spalte 'Technologie' (siehe
    bisherige Logik in build_consolidated_report) und registriert die
    Format-Spalten in fmt - analog zu _build_single_sheet(), aber fuer den
    Mehrfach-Technologie-Fall.
    """
    if not per_technology_source_dfs:
        return

    parts = []
    for slug, source_df in per_technology_source_dfs.items():
        part = source_df.copy()
        part["Technologie"] = per_technology_dfs[slug]["Technologie"].iloc[0]
        parts.append(part)

    combined = pd.concat(parts, ignore_index=True, sort=False)
    available_columns = ["Technologie"] + [
        col for col in spec.columns if col in combined.columns
    ]
    _write_sheet(combined[available_columns], writer, spec.sheet_name)

    fmt.decimal_columns += [col for col in spec.decimal_columns if col in available_columns]
    fmt.currency_columns += [col for col in spec.currency_columns if col in available_columns]
    fmt.percent_columns += [col for col in spec.percent_columns if col in available_columns]


# SheetSpec-Instanzen je 'einfachem' Reiter (siehe Konstanten weiter oben im
# Modul fuer die jeweiligen Spaltenlisten). Eine Stelle, an der die Reiter-
# Reihenfolge UND die Format-Zuordnung gemeinsam ersichtlich sind - bisher
# auf build_report()/build_consolidated_report() sowie mehrere Listen
# verteilt.
COVERAGE_SHEET_SPEC = SheetSpec(
    sheet_name=COVERAGE_SHEET_NAME,
    columns=COVERAGE_COLUMNS,
    decimal_columns=COVERAGE_DECIMAL_COLUMNS,
)

LEAD_TIME_SHEET_SPEC = SheetSpec(
    sheet_name=LEAD_TIME_SHEET_NAME,
    columns=LEAD_TIME_COLUMNS,
    decimal_columns=LEAD_TIME_DECIMAL_COLUMNS,
    currency_columns=LEAD_TIME_CURRENCY_COLUMNS,
)

SIMULATION_SHEET_SPEC = SheetSpec(
    sheet_name=SIMULATION_SHEET_NAME,
    columns=SIMULATION_COLUMNS,
    decimal_columns=SIMULATION_DECIMAL_COLUMNS,
    percent_columns=SIMULATION_PERCENT_COLUMNS,
)

SAFETY_STOCK_SHEET_SPEC = SheetSpec(
    sheet_name=SAFETY_STOCK_SHEET_NAME,
    columns=SAFETY_STOCK_COLUMNS,
    decimal_columns=SAFETY_STOCK_DECIMAL_COLUMNS,
    percent_columns=SAFETY_STOCK_PERCENT_COLUMNS,
)

WORKING_CAPITAL_SHEET_SPEC = SheetSpec(
    sheet_name=WORKING_CAPITAL_SHEET_NAME,
    columns=WORKING_CAPITAL_COLUMNS,
    decimal_columns=WORKING_CAPITAL_DECIMAL_COLUMNS,
    currency_columns=WORKING_CAPITAL_CURRENCY_COLUMNS,
    percent_columns=WORKING_CAPITAL_PERCENT_COLUMNS,
)

FORECAST_SHEET_SPEC = SheetSpec(
    sheet_name=FORECAST_SHEET_NAME,
    columns=FORECAST_COLUMNS,
    decimal_columns=FORECAST_DECIMAL_COLUMNS,
)

EXECUTIVE_SUMMARY_SHEET_SPEC = SheetSpec(
    sheet_name=EXECUTIVE_SUMMARY_SHEET_NAME,
    columns=EXECUTIVE_SUMMARY_COLUMNS,
    decimal_columns=EXECUTIVE_SUMMARY_DECIMAL_COLUMNS,
    currency_columns=EXECUTIVE_SUMMARY_CURRENCY_COLUMNS,
    percent_columns=EXECUTIVE_SUMMARY_PERCENT_COLUMNS,
)


def _build_trend_sheet(
    trend_df: pd.DataFrame | None,
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> tuple[list[str], list[str]]:
    """Schreibt den Reiter TREND_SHEET_NAME (Trend_MVER), sofern trend_df
    gesetzt ist. Bewusst NICHT Teil des generischen SheetSpec-Mechanismus
    (siehe Modul-Kommentar oben) - dynamische Monatsspalten erfordern eine
    eigene Spaltenzusammenstellung. Identisch zur bisherigen Logik in
    build_report(), nur ausgelagert.

    Returns:
        (trend_month_columns, trend_decimal_columns) - werden vom Aufrufer
        an _apply_formatting() weitergereicht (Outlier-Highlighting braucht
        trend_month_columns als Spalten-Lookup).
    """
    if trend_df is None:
        return [], []

    trend_month_columns = [
        col for col in trend_df.columns
        if col not in ("Material",) and col not in TREND_METRIC_COLUMNS
    ]
    trend_columns = ["Material"] + trend_month_columns + [
        col for col in TREND_METRIC_COLUMNS if col in trend_df.columns
    ]
    _write_sheet(trend_df[trend_columns], writer, TREND_SHEET_NAME)
    trend_decimal_columns = [
        col for col in TREND_DECIMAL_COLUMNS if col in trend_columns
    ] + trend_month_columns
    fmt.decimal_columns += trend_decimal_columns
    return trend_month_columns, trend_decimal_columns


def _build_trend_details_sheet(
    details_df: pd.DataFrame | None,
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> tuple[list[str], list[str], list[str]]:
    """Schreibt den Reiter TREND_DETAILS_SHEET_NAME (Trend_Details), sofern
    details_df gesetzt ist. Bewusst NICHT Teil des generischen SheetSpec-
    Mechanismus (siehe Modul-Kommentar oben) - Z-Score-Paarspalten je Monat
    erfordern eine eigene Spaltenzusammenstellung. Identisch zur bisherigen
    Logik in build_report(), nur ausgelagert.

    Returns:
        (details_value_columns, details_zscore_columns, details_metric_columns)
        - werden vom Aufrufer an _apply_formatting() weitergereicht.
    """
    if details_df is None:
        return [], [], []

    details_value_columns = [
        col for col in details_df.columns
        if col not in ("Material",)
        and col not in TREND_DETAILS_METRIC_COLUMNS
        and not col.endswith("_ZScore")
    ]
    details_zscore_columns = [f"{col}_ZScore" for col in details_value_columns]
    details_metric_columns = [
        col for col in TREND_DETAILS_METRIC_COLUMNS if col in details_df.columns
    ]
    details_columns = ["Material"] + details_metric_columns
    for value_col, zscore_col in zip(details_value_columns, details_zscore_columns):
        details_columns.append(value_col)
        if zscore_col in details_df.columns:
            details_columns.append(zscore_col)
    _write_sheet(details_df[details_columns], writer, TREND_DETAILS_SHEET_NAME)
    fmt.decimal_columns += details_value_columns + details_zscore_columns + details_metric_columns
    return details_value_columns, details_zscore_columns, details_metric_columns


def _build_executive_summary_sheet(
    summary_df: pd.DataFrame | None,
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> None:
    """Schreibt den Reiter EXECUTIVE_SUMMARY_SHEET_NAME (Einzelreport),
    sofern summary_df gesetzt ist. Nutzt den generischen SheetSpec-
    Mechanismus (siehe _build_single_sheet()) - Executive_Summary hat eine
    statische Spaltenliste, kein dynamisches Layout.
    """
    _build_single_sheet(EXECUTIVE_SUMMARY_SHEET_SPEC, summary_df, writer, fmt)


def _apply_priority_highlighting(ws, header_values: list) -> None:
    """Markiert die Zellen der Spalte 'Prioritaet' farblich nach Ampel-Logik
    (siehe PRIORITY_FILL_COLORS) - eigene, kleine Funktion statt Erweiterung
    von _apply_formatting() (siehe Modul-Kommentar bei PRIORITY_FILL_COLORS).
    Tut nichts, falls die Spalte 'Prioritaet' auf diesem Reiter nicht
    existiert (z.B. ein anderer Reiter als Executive_Summary).
    """
    if "Prioritaet" not in header_values:
        return
    col_idx = header_values.index("Prioritaet") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = PRIORITY_FILL_COLORS.get(cell.value)
        if fill is not None:
            cell.fill = fill


def _aggregate_for_dashboard(summary_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Berechnet die kleinen Aggregations-/Ranking-Tabellen, die als
    Datenquelle fuer die Dashboard-Charts dienen (siehe
    _build_dashboard_sheet()). Eigenstaendige Funktion, damit dieselbe
    Aggregationslogik fuer Einzel- UND konsolidierten Report verwendet wird
    (build_consolidated_report ruft sie auf dem zusammengefuehrten
    Executive-Summary-DataFrame ueber alle Technologien auf).

    Args:
        summary_df: Ergebnis von executive_summary.build_executive_summary(),
            ggf. zusaetzlich um 'Technologie' ergaenzt (konsolidierter Fall).

    Returns:
        Dict mit folgenden Eintraegen (jeweils ein DataFrame, leer falls
        die zugrunde liegende Spalte fehlt oder keine gueltigen Werte hat):
            - 'bestandswert_je_technologie': Technologie, Bestandswert_EUR
              (nur im konsolidierten Fall sinnvoll mit >1 Zeile - im
              Einzelreport entsteht hier eine einzeilige Tabelle, siehe
              _build_dashboard_sheet() fuer den Umgang damit)
            - 'prioritaet_verteilung': Prioritaet, Anzahl
            - 'trend_verteilung': Trend_Einstufung, Anzahl
            - 'top_risiko_materialien': Material, Materialkurztext,
              Fehlmengen_Wahrscheinlichkeit_Horizont (Top 15, absteigend
              sortiert, NUR Zeilen mit vorhandenem Wert)
            - 'wert_vs_reichweite': Material, Bestandswert_EUR,
              Reichweite_Monate, ABC-Kennzeichen (fuer Scatter/Risikomatrix,
              NUR Zeilen mit beiden Werten vorhanden)
    """
    result: dict[str, pd.DataFrame] = {}

    if "Technologie" in summary_df.columns:
        result["bestandswert_je_technologie"] = (
            summary_df.groupby("Technologie", as_index=False)["Bestandswert_EUR"]
            .sum()
            .sort_values("Bestandswert_EUR", ascending=False)
        )
    else:
        result["bestandswert_je_technologie"] = pd.DataFrame(
            columns=["Technologie", "Bestandswert_EUR"]
        )

    result["prioritaet_verteilung"] = (
        summary_df["Prioritaet"].value_counts().rename_axis("Prioritaet")
        .reset_index(name="Anzahl")
    )

    if "Trend_Einstufung" in summary_df.columns:
        trend_counts = summary_df["Trend_Einstufung"].dropna()
        result["trend_verteilung"] = (
            trend_counts.value_counts().rename_axis("Trend_Einstufung")
            .reset_index(name="Anzahl")
        )
    else:
        result["trend_verteilung"] = pd.DataFrame(columns=["Trend_Einstufung", "Anzahl"])

    if "Fehlmengen_Wahrscheinlichkeit_Horizont" in summary_df.columns:
        top_risk = summary_df.dropna(subset=["Fehlmengen_Wahrscheinlichkeit_Horizont"])
        top_risk = top_risk.sort_values(
            "Fehlmengen_Wahrscheinlichkeit_Horizont", ascending=False
        ).head(15)
        result["top_risiko_materialien"] = top_risk[
            ["Material", "Materialkurztext", "Fehlmengen_Wahrscheinlichkeit_Horizont"]
        ]
    else:
        result["top_risiko_materialien"] = pd.DataFrame(
            columns=["Material", "Materialkurztext", "Fehlmengen_Wahrscheinlichkeit_Horizont"]
        )

    if "Reichweite_Monate" in summary_df.columns:
        scatter = summary_df.dropna(subset=["Bestandswert_EUR", "Reichweite_Monate"])
        result["wert_vs_reichweite"] = scatter[
            ["Material", "Bestandswert_EUR", "Reichweite_Monate", "ABC-Kennzeichen"]
        ]
    else:
        result["wert_vs_reichweite"] = pd.DataFrame(
            columns=["Material", "Bestandswert_EUR", "Reichweite_Monate", "ABC-Kennzeichen"]
        )

    return result


def _write_dashboard_helper_table(
    ws, df: pd.DataFrame, start_row: int, start_col: int
) -> tuple[int, int]:
    """Schreibt eine kleine Hilfstabelle (Header + Zeilen) ab der
    angegebenen Zelle direkt auf den Dashboard-Reiter (siehe Modul-Kommentar
    bei DASHBOARD_SHEET_NAME - openpyxl-Charts brauchen ihre Datenquelle im
    selben Workbook). Leere DataFrames werden trotzdem mit Header
    geschrieben (0 Datenzeilen), damit nachfolgende Charts eine gueltige,
    wenn auch leere Reference erhalten, statt den Lauf abzubrechen.

    Returns:
        (end_row, end_col): letzte beschriebene Zeile/Spalte (1-indiziert),
        damit der Aufrufer die Chart-Reference exakt darauf aufbauen kann.
    """
    for col_offset, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + col_offset, value=col_name)
    for row_offset, (_, row) in enumerate(df.iterrows(), start=1):
        for col_offset, col_name in enumerate(df.columns):
            ws.cell(
                row=start_row + row_offset,
                column=start_col + col_offset,
                value=row[col_name],
            )
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    return end_row, end_col


def _build_dashboard_sheet(
    summary_df: pd.DataFrame | None,
    writer: pd.ExcelWriter,
) -> None:
    """Schreibt den Reiter DASHBOARD_SHEET_NAME mit nativen Excel-Diagrammen
    (openpyxl.chart), sofern summary_df gesetzt ist. Tut nichts, falls
    summary_df None ist - haelt den Aufruf abwaertskompatibel, analog zu
    allen anderen *_df-Parametern in build_report()/build_consolidated_
    report().

    Vier Diagramme (Klaerung mit Max, 29.06.2026, siehe Projektstatus.md
    Abschnitt 4f):
        1. Bestandswert je Technologie (Balken) - nur im konsolidierten
           Report mit >1 Technologie aussagekraeftig; im Einzelreport
           entsteht eine einzeilige Tabelle, das Chart wird trotzdem
           erzeugt (zeigt dann nur den einen Wert), bewusst KEIN
           Sonderfall im Code, um die Funktion fuer beide Aufrufer
           identisch zu halten.
        2. Prioritaets-Verteilung (Kreisdiagramm) - aus _classify_priority()
           in executive_summary.py.
        3. Top-15-Fehlmengen-Risiko (Balken, horizontal waere lesbarer,
           openpyxl unterstuetzt dies nicht direkt ueber BarChart.type,
           daher vertikaler Balken mit Material-Beschriftung).
        4. Bestandswert vs. Reichweite (Scatter, Datenpunkte nach
           ABC-Kennzeichen NICHT separat eingefaerbt - openpyxl-
           ScatterChart unterstuetzt keine kategoriale Farbskala ohne
           mehrere Series; eine Series pro ABC-Klasse waere moeglich,
           wird hier bewusst auf eine einzelne Series reduziert, um die
           Komplexitaet gering zu halten - ABC-Kennzeichen bleibt als
           Spalte in der Hilfstabelle sichtbar, falls Max das per Hand
           in Excel weiter aufteilen will).

    Die Diagramme verwenden Hilfstabellen, die auf demselben Reiter
    geschrieben werden (siehe _write_dashboard_helper_table()) - eine
    Anordnung in zwei Spalten-Bloecken (Diagramme links ab Spalte A,
    Hilfstabellen rechts ab Spalte N), damit die Diagramme beim Oeffnen
    der Datei direkt sichtbar sind, ohne dass die Rohdaten dahinter stoeren.
    """
    if summary_df is None:
        return

    aggregates = _aggregate_for_dashboard(summary_df)

    workbook = writer.book
    ws = workbook.create_sheet(DASHBOARD_SHEET_NAME)
    ws.cell(row=1, column=1, value="GMTS Dashboard")
    ws["A1"].font = Font(bold=True, size=14)

    helper_col = 14  # Spalte N - Hilfstabellen rechts neben den Diagrammen
    row = DASHBOARD_HELPER_TABLE_START_ROW

    # --- 1. Bestandswert je Technologie (Balken) ---------------------------
    bwt_df = aggregates["bestandswert_je_technologie"]
    bwt_start_row = row
    end_row, end_col = _write_dashboard_helper_table(ws, bwt_df, bwt_start_row, helper_col)
    if len(bwt_df) > 0:
        chart = BarChart()
        chart.title = "Bestandswert je Technologie (EUR)"
        chart.y_axis.title = "Bestandswert (EUR)"
        chart.x_axis.title = "Technologie"
        data_ref = Reference(
            ws, min_col=helper_col + 1, min_row=bwt_start_row, max_row=end_row
        )
        cat_ref = Reference(
            ws, min_col=helper_col, min_row=bwt_start_row + 1, max_row=end_row
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.width, chart.height = 16, 9
        ws.add_chart(chart, "A2")
    row = end_row + 3

    # --- 2. Prioritaets-Verteilung (Kreisdiagramm) -------------------------
    prio_df = aggregates["prioritaet_verteilung"]
    prio_start_row = row
    end_row, end_col = _write_dashboard_helper_table(ws, prio_df, prio_start_row, helper_col)
    if len(prio_df) > 0:
        chart = PieChart()
        chart.title = "Prioritaets-Verteilung"
        data_ref = Reference(
            ws, min_col=helper_col + 1, min_row=prio_start_row, max_row=end_row
        )
        cat_ref = Reference(
            ws, min_col=helper_col, min_row=prio_start_row + 1, max_row=end_row
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.width, chart.height = 12, 9
        ws.add_chart(chart, "A20")
    row = end_row + 3

    # --- 3. Top-15-Fehlmengen-Risiko (Balken) ------------------------------
    risk_df = aggregates["top_risiko_materialien"]
    risk_start_row = row
    end_row, end_col = _write_dashboard_helper_table(ws, risk_df, risk_start_row, helper_col)
    if len(risk_df) > 0:
        chart = BarChart()
        chart.title = "Top-Materialien nach Fehlmengen-Wahrscheinlichkeit"
        chart.y_axis.title = "Fehlmengen-Wahrscheinlichkeit (Horizont)"
        chart.x_axis.title = "Material"
        # Spaltenreihenfolge in risk_df: Material, Materialkurztext,
        # Fehlmengen_Wahrscheinlichkeit_Horizont -> dritte Spalte = Werte.
        value_col_idx = helper_col + 2
        data_ref = Reference(
            ws, min_col=value_col_idx, min_row=risk_start_row, max_row=end_row
        )
        cat_ref = Reference(
            ws, min_col=helper_col, min_row=risk_start_row + 1, max_row=end_row
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.width, chart.height = 20, 10
        ws.add_chart(chart, "A38")
    row = end_row + 3

    # --- 4. Bestandswert vs. Reichweite (Scatter, ABC/XYZ-Risikomatrix) ---
    scatter_df = aggregates["wert_vs_reichweite"]
    scatter_start_row = row
    end_row, end_col = _write_dashboard_helper_table(
        ws, scatter_df, scatter_start_row, helper_col
    )
    if len(scatter_df) > 0:
        chart = ScatterChart()
        chart.title = "Bestandswert vs. Reichweite (Risikomatrix)"
        chart.x_axis.title = "Reichweite (Monate)"
        chart.y_axis.title = "Bestandswert (EUR)"
        # Spaltenreihenfolge in scatter_df: Material, Bestandswert_EUR,
        # Reichweite_Monate, ABC-Kennzeichen.
        x_ref = Reference(
            ws, min_col=helper_col + 2, min_row=scatter_start_row + 1, max_row=end_row
        )
        y_ref = Reference(
            ws, min_col=helper_col + 1, min_row=scatter_start_row, max_row=end_row
        )
        series = Series(y_ref, x_ref, title_from_data=True)
        series.marker.symbol = "circle"
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        chart.width, chart.height = 18, 11
        ws.add_chart(chart, "A60")


def build_report(
    df: pd.DataFrame,
    cfg: RunConfig,
    coverage_df: pd.DataFrame | None = None,
    trend_df: pd.DataFrame | None = None,
    outlier_cells: dict[str, list[str]] | None = None,
    details_df: pd.DataFrame | None = None,
    lead_time_df: pd.DataFrame | None = None,
    simulation_df: pd.DataFrame | None = None,
    safety_stock_df: pd.DataFrame | None = None,
    forecast_df: pd.DataFrame | None = None,
    working_capital_df: pd.DataFrame | None = None,
    summary_df: pd.DataFrame | None = None,
) -> None:
    """Baut den kombinierten Excel-Report fuer den ZMLAG-Lauf und speichert
    ihn unter cfg.report_output_path.

    Erzeugt immer den Reiter 'ZMLAG_Bestand' (Materialstammdaten,
    Stichtags-Werte, Kennzahlen aus stock_analysis.py). Jede weitere Phase
    ergaenzt - sofern die jeweilige *_df uebergeben wird - einen eigenen
    Reiter in derselben Arbeitsmappe (siehe SheetSpec-Konstanten und
    Modul-Kommentar "Generischer Section-Mechanismus" oben):
        coverage_df       -> Reiter 'Reichweite'      (Phase 2)
        trend_df          -> Reiter 'Trend_MVER'      (Phase 3)
        details_df        -> Reiter 'Trend_Details'   (Transparenz-Ergaenzung)
        lead_time_df      -> Reiter 'Wiederbeschaffung' (Phase 4)
        simulation_df     -> Reiter 'Simulation_Verbrauch' (Baustein A)
        safety_stock_df   -> Reiter 'Fruehwarnsystem' (Phase 5)
        forecast_df       -> Reiter 'Forecast_MVER'   (Modellauswahl-Forecast,
                             29.06.2026 - eigenstaendig neben trend_df/
                             Trend_MVER, siehe forecast_analysis.py)
        working_capital_df -> Reiter 'Working_Capital' (Phase 6 Fortsetzung,
                             30.06.2026 - siehe working_capital.py und
                             Projektstatus.md Abschnitt 4i)
        summary_df        -> Reiter 'Executive_Summary' PLUS Reiter
                             'Dashboard' (Phase 6, 29.06.2026 - siehe
                             executive_summary.py; Dashboard wird IMMER
                             zusaetzlich erzeugt, sobald summary_df gesetzt
                             ist, da die Diagramme direkt auf summary_df
                             aufbauen, kein eigener Parameter dafuer noetig)
    Jeder *_df-Parameter ist unabhaengig optional (Default None) und haelt
    den Aufruf damit abwaertskompatibel fuer Laeufe, die einzelne Phasen
    (noch) nicht nutzen.

    Args:
        df: vollstaendig angereicherter DataFrame (nach calculate_stock_development
            und classify_risk_by_abc_xyz aus stock_analysis.py)
        cfg: aktive RunConfig
        coverage_df: optionales Ergebnis von coverage_analysis.build_coverage_table().
        trend_df: optionales Ergebnis von trend_analysis.build_trend_table()
            (erster Tupel-Eintrag).
        outlier_cells: optionales Ergebnis von trend_analysis.build_trend_table()
            (zweiter Tupel-Eintrag) - Dict {Material: [Monatsspalten-Namen]},
            genutzt fuer die Markierung in Trend_MVER UND Trend_Details.
        details_df: optionales Ergebnis von trend_analysis.build_trend_table()
            (dritter Tupel-Eintrag) - unabhaengig von trend_df steuerbar.
        lead_time_df: optionales Ergebnis von dispo_abcxyz_loader.
            build_lead_time_table(), bereits auf die Materialien dieser
            Technologie gefiltert.
        simulation_df: optionales Ergebnis von simulation_analysis.
            build_simulation_table().
        safety_stock_df: optionales Ergebnis von safety_stock.
            build_safety_stock_table() (Phase 5, ab 29.06.2026).
        forecast_df: optionales Ergebnis von forecast_analysis.
            build_forecast_table() (ab 29.06.2026) - eigenstaendig neben
            trend_df, KEINE Abhaengigkeit zwischen beiden in diesem Modul.
        working_capital_df: optionales Ergebnis von working_capital.
            build_working_capital_table() (Phase 6 Fortsetzung, ab
            30.06.2026) - siehe Projektstatus.md Abschnitt 4i.
        summary_df: optionales Ergebnis von executive_summary.
            build_executive_summary() (Phase 6, ab 29.06.2026). Erzeugt
            BEIDE neuen Reiter (Executive_Summary, Dashboard) - siehe
            _build_executive_summary_sheet()/_build_dashboard_sheet().
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    period_columns = get_period_columns(df)
    zmlag_columns = ZMLAG_BASE_COLUMNS + period_columns + ZMLAG_CALCULATED_COLUMNS

    fmt = _FormatCollector(
        currency_columns=CURRENCY_COLUMNS_FIXED + period_columns,
        percent_columns=list(PERCENT_COLUMNS),
    )

    with pd.ExcelWriter(cfg.report_output_path, engine="openpyxl") as writer:
        _write_sheet(df[zmlag_columns], writer, ZMLAG_SHEET_NAME)

        _build_single_sheet(COVERAGE_SHEET_SPEC, coverage_df, writer, fmt)
        trend_month_columns, trend_decimal_columns = _build_trend_sheet(trend_df, writer, fmt)
        details_value_columns, details_zscore_columns, details_metric_columns = (
            _build_trend_details_sheet(details_df, writer, fmt)
        )
        _build_single_sheet(LEAD_TIME_SHEET_SPEC, lead_time_df, writer, fmt)
        _build_single_sheet(SIMULATION_SHEET_SPEC, simulation_df, writer, fmt)
        _build_single_sheet(SAFETY_STOCK_SHEET_SPEC, safety_stock_df, writer, fmt)
        _build_single_sheet(FORECAST_SHEET_SPEC, forecast_df, writer, fmt)
        _build_single_sheet(WORKING_CAPITAL_SHEET_SPEC, working_capital_df, writer, fmt)
        _build_executive_summary_sheet(summary_df, writer, fmt)
        _build_dashboard_sheet(summary_df, writer)

    _apply_formatting(
        cfg.report_output_path,
        fmt.currency_columns,
        fmt.percent_columns,
        decimal_columns=fmt.decimal_columns,
        trend_decimal_columns=trend_decimal_columns,
        outlier_cells=outlier_cells if trend_df is not None else None,
        trend_month_columns=trend_month_columns,
        details_value_columns=details_value_columns,
        details_zscore_columns=details_zscore_columns,
        details_outlier_cells=outlier_cells if details_df is not None else None,
        details_metric_columns=details_metric_columns,
    )


def _build_consolidated_trend_sheet(
    per_technology_trend_dfs: dict[str, pd.DataFrame] | None,
    per_technology_dfs: dict[str, pd.DataFrame],
    per_technology_outlier_cells: dict[str, dict[str, list[str]]] | None,
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> tuple[list[str], list[str], dict[str, list[str]]]:
    """Konsolidierte Variante von _build_trend_sheet() (siehe dort fuer den
    Grund, warum Trend_MVER nicht im generischen SheetSpec-Mechanismus
    steckt). Identisch zur bisherigen Logik in build_consolidated_report(),
    nur ausgelagert.

    Returns:
        (trend_month_columns, trend_decimal_columns, combined_outlier_cells)
    """
    combined_outlier_cells: dict[str, list[str]] = {}
    if not per_technology_trend_dfs:
        return [], [], combined_outlier_cells

    trend_parts = []
    for slug, trend_df in per_technology_trend_dfs.items():
        part = trend_df.copy()
        part["Technologie"] = per_technology_dfs[slug]["Technologie"].iloc[0]
        trend_parts.append(part)
        # outlier_cells ist je Technologie-Slug auf Material-Ebene eindeutig
        # (ein Material gehoert genau einer Technologie an) - einfaches
        # Zusammenfuehren der Dicts ist daher unproblematisch.
        if per_technology_outlier_cells and slug in per_technology_outlier_cells:
            combined_outlier_cells.update(per_technology_outlier_cells[slug])

    combined_trend = pd.concat(trend_parts, ignore_index=True, sort=False)

    trend_month_columns = sorted(
        {
            col
            for trend_df in per_technology_trend_dfs.values()
            for col in trend_df.columns
            if col not in ("Material",) and col not in TREND_METRIC_COLUMNS
        }
    )
    for col in trend_month_columns:
        if col not in combined_trend.columns:
            combined_trend[col] = pd.NA

    trend_columns = ["Technologie", "Material"] + trend_month_columns + [
        col for col in TREND_METRIC_COLUMNS if col in combined_trend.columns
    ]
    _write_sheet(combined_trend[trend_columns], writer, TREND_SHEET_NAME)
    trend_decimal_columns = [
        col for col in TREND_DECIMAL_COLUMNS if col in trend_columns
    ] + trend_month_columns
    fmt.decimal_columns += trend_decimal_columns
    return trend_month_columns, trend_decimal_columns, combined_outlier_cells


def _build_consolidated_trend_details_sheet(
    per_technology_details_dfs: dict[str, pd.DataFrame] | None,
    per_technology_dfs: dict[str, pd.DataFrame],
    writer: pd.ExcelWriter,
    fmt: _FormatCollector,
) -> tuple[list[str], list[str], list[str], bool]:
    """Konsolidierte Variante von _build_trend_details_sheet() (siehe dort
    fuer den Grund, warum Trend_Details nicht im generischen SheetSpec-
    Mechanismus steckt). Identisch zur bisherigen Logik in
    build_consolidated_report(), nur ausgelagert.

    Returns:
        (details_value_columns, details_zscore_columns, details_metric_columns,
        was_written) - was_written zeigt an, ob der Reiter erzeugt wurde
        (fuer die details_outlier_cells-Weitergabe an _apply_formatting()).
    """
    if not per_technology_details_dfs:
        return [], [], [], False

    details_parts = []
    for slug, details_df in per_technology_details_dfs.items():
        part = details_df.copy()
        part["Technologie"] = per_technology_dfs[slug]["Technologie"].iloc[0]
        details_parts.append(part)

    combined_details = pd.concat(details_parts, ignore_index=True, sort=False)

    details_value_columns = sorted(
        {
            col
            for details_df in per_technology_details_dfs.values()
            for col in details_df.columns
            if col not in ("Material",)
            and col not in TREND_DETAILS_METRIC_COLUMNS
            and not col.endswith("_ZScore")
        }
    )
    details_zscore_columns = [f"{col}_ZScore" for col in details_value_columns]
    for col in details_value_columns + details_zscore_columns:
        if col not in combined_details.columns:
            combined_details[col] = pd.NA

    details_metric_columns = [
        col for col in TREND_DETAILS_METRIC_COLUMNS if col in combined_details.columns
    ]
    details_columns = ["Technologie", "Material"] + details_metric_columns
    for value_col, zscore_col in zip(details_value_columns, details_zscore_columns):
        details_columns.append(value_col)
        details_columns.append(zscore_col)
    _write_sheet(combined_details[details_columns], writer, TREND_DETAILS_SHEET_NAME)
    fmt.decimal_columns += details_value_columns + details_zscore_columns + details_metric_columns
    return details_value_columns, details_zscore_columns, details_metric_columns, True


def build_consolidated_report(
    per_technology_dfs: dict[str, pd.DataFrame],
    output_path,
    per_technology_coverage_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_trend_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_outlier_cells: dict[str, dict[str, list[str]]] | None = None,
    per_technology_details_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_lead_time_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_simulation_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_safety_stock_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_forecast_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_working_capital_dfs: dict[str, pd.DataFrame] | None = None,
    per_technology_summary_dfs: dict[str, pd.DataFrame] | None = None,
) -> None:
    """Baut einen konsolidierten Excel-Report mit allen Technologien in EINER
    gemeinsamen Tabelle (Reiter 'Alle_Technologien'), zusaetzlich zu den
    weiterhin separat erzeugten Einzelreports je Technologie (build_report).

    Hintergrund (Multi-Technologie-Lauf, siehe main.py): Wenn alle 8
    Technologien in einem Durchlauf verarbeitet werden, soll neben den
    gewohnten 8 Einzelreports auch ein Gesamtueberblick entstehen. Die
    Spalte 'Technologie' (materialgenau aus SAP, siehe data_loader.py)
    erlaubt es, in dieser gemeinsamen Tabelle nach Technologie zu filtern
    oder zu pivotieren, ohne dass dafuer mehrere Reiter noetig sind.

    Da sich die Stichtags-Spalten (JJJJMM) zwischen Technologien durch
    unterschiedliche Exportzeitpunkte unterscheiden KOENNEN, wird hier die
    Vereinigung aller in irgendeiner Technologie vorkommenden Stichtags-
    Spalten verwendet; fehlende Werte je Material bleiben leer (NaN).

    Jede weitere Phase ergaenzt - sofern das jeweilige per_technology_*_dfs
    uebergeben wird - einen technologieuebergreifenden Reiter, analog zu
    build_report() aber mit allen Technologien in EINER Tabelle (ergaenzt
    um die Spalte 'Technologie'). Siehe SheetSpec-Konstanten und Modul-
    Kommentar "Generischer Section-Mechanismus" oben.

    Args:
        per_technology_dfs: Mapping Technologie-Slug -> vollstaendig
            angereicherter DataFrame (nach calculate_stock_development und
            classify_risk_by_abc_xyz), wie er auch an build_report() je
            Technologie uebergeben wird.
        output_path: Zielpfad fuer die konsolidierte Excel-Datei.
        per_technology_coverage_dfs: optionales Mapping Technologie-Slug ->
            coverage_df. Falls None/leer, wird KEIN konsolidierter
            Reichweite-Reiter erzeugt.
        per_technology_trend_dfs: optionales Mapping Technologie-Slug ->
            trend_df (erster Tupel-Eintrag). Falls None/leer, wird KEIN
            konsolidierter Trend_MVER-Reiter erzeugt.
        per_technology_outlier_cells: optionales Mapping Technologie-Slug ->
            outlier_cells (zweiter Tupel-Eintrag), fuer die farbliche
            Markierung im konsolidierten Reiter.
        per_technology_details_dfs: optionales Mapping Technologie-Slug ->
            details_df (dritter Tupel-Eintrag). Falls None/leer, wird KEIN
            konsolidierter Trend_Details-Reiter erzeugt.
        per_technology_lead_time_dfs: optionales Mapping Technologie-Slug ->
            lead_time_df. Falls None/leer, wird KEIN konsolidierter
            Wiederbeschaffung-Reiter erzeugt.
        per_technology_simulation_dfs: optionales Mapping Technologie-Slug ->
            simulation_df. Falls None/leer, wird KEIN konsolidierter
            Simulation_Verbrauch-Reiter erzeugt.
        per_technology_safety_stock_dfs: optionales Mapping Technologie-Slug ->
            safety_stock_df (Phase 5, ab 29.06.2026). Falls None/leer, wird
            KEIN konsolidierter Fruehwarnsystem-Reiter erzeugt.
        per_technology_forecast_dfs: optionales Mapping Technologie-Slug ->
            forecast_df (Modellauswahl-Forecast, ab 29.06.2026, siehe
            forecast_analysis.py). Falls None/leer, wird KEIN konsolidierter
            Forecast_MVER-Reiter erzeugt.
        per_technology_working_capital_dfs: optionales Mapping Technologie-
            Slug -> working_capital_df (Phase 6 Fortsetzung, ab 30.06.2026,
            siehe working_capital.py). Falls None/leer, wird KEIN
            konsolidierter Working_Capital-Reiter erzeugt.
        per_technology_summary_dfs: optionales Mapping Technologie-Slug ->
            summary_df (Ergebnis von executive_summary.
            build_executive_summary(), Phase 6, ab 29.06.2026). Falls
            None/leer, werden KEIN konsolidierter Executive_Summary- UND
            KEIN Dashboard-Reiter erzeugt (Dashboard haengt direkt an der
            zusammengefuehrten Summary, siehe _build_dashboard_sheet()).
    """
    combined = pd.concat(per_technology_dfs.values(), ignore_index=True, sort=False)

    all_period_columns = sorted(
        {col for df in per_technology_dfs.values() for col in get_period_columns(df)},
        key=str,
    )
    consolidated_columns = (
        ZMLAG_BASE_COLUMNS + all_period_columns + ZMLAG_CALCULATED_COLUMNS
    )
    # Nicht jede Technologie hat zwingend dieselben Stichtags-Spalten -
    # fehlende Spalten je Zeile werden hier als NaN ergaenzt, statt einen
    # KeyError auszuloesen.
    for col in consolidated_columns:
        if col not in combined.columns:
            combined[col] = pd.NA

    fmt = _FormatCollector(
        currency_columns=CURRENCY_COLUMNS_FIXED + all_period_columns,
        percent_columns=list(PERCENT_COLUMNS),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(combined[consolidated_columns], writer, CONSOLIDATED_SHEET_NAME)

        _build_consolidated_sheet(
            COVERAGE_SHEET_SPEC, per_technology_coverage_dfs, per_technology_dfs, writer, fmt
        )
        trend_month_columns, trend_decimal_columns, combined_outlier_cells = (
            _build_consolidated_trend_sheet(
                per_technology_trend_dfs, per_technology_dfs, per_technology_outlier_cells, writer, fmt
            )
        )
        details_value_columns, details_zscore_columns, details_metric_columns, details_written = (
            _build_consolidated_trend_details_sheet(
                per_technology_details_dfs, per_technology_dfs, writer, fmt
            )
        )
        _build_consolidated_sheet(
            LEAD_TIME_SHEET_SPEC, per_technology_lead_time_dfs, per_technology_dfs, writer, fmt
        )
        _build_consolidated_sheet(
            SIMULATION_SHEET_SPEC, per_technology_simulation_dfs, per_technology_dfs, writer, fmt
        )
        _build_consolidated_sheet(
            SAFETY_STOCK_SHEET_SPEC, per_technology_safety_stock_dfs, per_technology_dfs, writer, fmt
        )
        _build_consolidated_sheet(
            FORECAST_SHEET_SPEC, per_technology_forecast_dfs, per_technology_dfs, writer, fmt
        )
        _build_consolidated_sheet(
            WORKING_CAPITAL_SHEET_SPEC, per_technology_working_capital_dfs, per_technology_dfs, writer, fmt
        )

        combined_summary = None
        if per_technology_summary_dfs:
            summary_parts = []
            for slug, summary_df in per_technology_summary_dfs.items():
                part = summary_df.copy()
                part["Technologie"] = per_technology_dfs[slug]["Technologie"].iloc[0]
                summary_parts.append(part)
            combined_summary = pd.concat(summary_parts, ignore_index=True, sort=False)

            summary_available_columns = ["Technologie"] + [
                col for col in EXECUTIVE_SUMMARY_COLUMNS if col in combined_summary.columns
            ]
            _write_sheet(
                combined_summary[summary_available_columns], writer, EXECUTIVE_SUMMARY_SHEET_NAME
            )
            fmt.decimal_columns += [
                col for col in EXECUTIVE_SUMMARY_DECIMAL_COLUMNS
                if col in summary_available_columns
            ]
            fmt.currency_columns += [
                col for col in EXECUTIVE_SUMMARY_CURRENCY_COLUMNS
                if col in summary_available_columns
            ]
            fmt.percent_columns += [
                col for col in EXECUTIVE_SUMMARY_PERCENT_COLUMNS
                if col in summary_available_columns
            ]

            _build_dashboard_sheet(combined_summary, writer)

    _apply_formatting(
        output_path,
        fmt.currency_columns,
        fmt.percent_columns,
        decimal_columns=fmt.decimal_columns,
        trend_decimal_columns=trend_decimal_columns,
        outlier_cells=combined_outlier_cells if trend_month_columns else None,
        trend_month_columns=trend_month_columns,
        details_value_columns=details_value_columns,
        details_zscore_columns=details_zscore_columns,
        details_outlier_cells=combined_outlier_cells if details_written else None,
        details_metric_columns=details_metric_columns,
    )


def _write_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, sheet_name: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def _apply_formatting(
    file_path,
    currency_columns: list[str],
    percent_columns: list[str],
    decimal_columns: list[str] | None = None,
    trend_decimal_columns: list[str] | None = None,
    outlier_cells: dict[str, list[str]] | None = None,
    trend_month_columns: list[str] | None = None,
    details_value_columns: list[str] | None = None,
    details_zscore_columns: list[str] | None = None,
    details_outlier_cells: dict[str, list[str]] | None = None,
    details_metric_columns: list[str] | None = None,
) -> None:
    """Wendet einheitliche Kopfzeilen-Formatierung, Spaltenbreiten, Text-
    Zellformat fuer Materialnummern sowie Waehrungs-, Prozent- und (seit
    Phase 2) einfache Dezimalformate auf alle Reiter des gespeicherten
    Workbooks an.

    Materialnummern bestehen ausschliesslich aus Ziffern (z.B. '00314814').
    Ohne explizites Text-Zellformat interpretiert Excel solche Zellen beim
    Oeffnen automatisch als Zahl und entfernt dabei fuehrende Nullen - das
    wuerde beim Abgleich mit anderen SAP-Exporten (MVER etc.) zu
    Fehlzuordnungen fuehren. Das Zellformat '@' (Text) verhindert das.

    Waehrungsspalten (Stichtags-Werte, Laufender Wert, Bestandswert_Start/
    Letzter_Stichtag/Ende, Veraenderung_Absolut) erhalten das Format
    CURRENCY_FORMAT, Prozentspalten (Veraenderung_Prozent) das Format
    PERCENT_FORMAT. Letzteres setzt voraus, dass der zugrunde liegende
    Rechenwert bereits als Dezimalzahl (z.B. -0.2907) vorliegt, siehe
    stock_analysis.calculate_stock_development.

    Dezimalspalten (Phase 2: Bestandsmenge, Ø_Verbrauch_12M, Reichweite in
    Monaten etc.; Phase 3: Trend-Steigung, Streuung, rohe Monatsspalten)
    erhalten COVERAGE_DECIMAL_FORMAT - bewusst KEIN Waehrungsformat, da es
    sich um Stueck/Monate handelt, kein Geld.

    Phase 3 (ab 25.06.2026): Als Sondereffekt erkannte Zellen in den rohen
    Monatsspalten des Trend_MVER-Reiters werden zusaetzlich mit
    OUTLIER_CELL_FILL farblich hervorgehoben (siehe outlier_cells) - der
    Zahlenwert selbst bleibt unveraendert, nur optisch markiert (Klaerung
    mit Max, 24.06.2026: Peak soll im Report erkennbar bleiben, auch wenn
    er fuer Trend/Saisonalitaet herausgerechnet wurde). Die Markierung wird
    NUR auf dem Reiter TREND_SHEET_NAME angewendet, ueber das Mapping
    Material (Zeile) x Monatsspalte (Spalte), unabhaengig von der
    Zeilenreihenfolge im DataFrame - daher wird je Reiter zunaechst die
    Material-Spalte ausgelesen, um pro Zeile das richtige Material zu
    bestimmen.

    Args:
        file_path: Pfad zur gespeicherten Excel-Datei
        currency_columns: Spaltennamen, die als Waehrung formatiert werden sollen
        percent_columns: Spaltennamen, die als Prozent formatiert werden sollen
        decimal_columns: Spaltennamen (Phase 2), die als einfache Dezimalzahl
            (2 Nachkommastellen, kein Waehrungssymbol) formatiert werden sollen
        trend_decimal_columns: Spaltennamen (Phase 3: Kennzahlen + rohe
            Monatsspalten), die ebenfalls COVERAGE_DECIMAL_FORMAT erhalten -
            separater Parameter nur zur klareren Aufrufstelle in build_report/
            build_consolidated_report, wird intern mit decimal_columns
            zusammengefuehrt.
        outlier_cells: Dict {Material: [Monatsspalten-Namen]} fuer die
            farbliche Markierung im Reiter TREND_SHEET_NAME. None/leer,
            falls kein Trend_MVER-Reiter erzeugt wurde.
        trend_month_columns: Spaltennamen der rohen Monatsspalten im
            Trend_MVER-Reiter - wird benoetigt, um beim Iterieren ueber die
            Spalten von ws.columns auf den richtigen Spaltennamen
            zurueckschliessen zu koennen (siehe Anwendung von outlier_cells
            weiter unten).
        details_value_columns: Spaltennamen der rohen Monats-Wertspalten im
            Reiter TREND_DETAILS_SHEET_NAME (ab 25.06.2026) - analog zu
            trend_month_columns, aber fuer den Transparenz-Reiter. Erhalten
            COVERAGE_DECIMAL_FORMAT wie die uebrigen Dezimalspalten.
        details_zscore_columns: zugehoerige '<Monat>_ZScore'-Spaltennamen im
            selben Reiter - erhalten ebenfalls COVERAGE_DECIMAL_FORMAT, da
            ein Z-Score eine reine Verhaeltniszahl ist (kein Stueck-/
            Geldwert), aber als Dezimalzahl trotzdem gut lesbar sein soll.
        details_outlier_cells: Dict {Material: [Monatsspalten-Namen]} fuer
            die farbliche Markierung im Reiter TREND_DETAILS_SHEET_NAME -
            inhaltlich identisch zu outlier_cells (dieselbe Sondereffekt-
            Erkennung), hier separat uebergeben, da sie auf einem anderen
            Reiter mit anderem Spaltenlayout angewendet wird. Die Markierung
            betrifft nur die Wertspalten, nicht die Z-Score-Spalten - der
            Z-Score selbst macht den Ausreisser bereits an seinem Betrag
            erkennbar, eine zusaetzliche Markierung dort waere redundant.
        details_metric_columns: Spaltennamen (Verbrauch_Median,
            Verbrauch_Streuung_MAD) im Reiter TREND_DETAILS_SHEET_NAME, die
            ebenfalls COVERAGE_DECIMAL_FORMAT erhalten.
    """
    from openpyxl import load_workbook

    details_value_columns = details_value_columns or []
    details_zscore_columns = details_zscore_columns or []
    details_metric_columns = details_metric_columns or []

    decimal_columns = (
        list(decimal_columns or [])
        + list(trend_decimal_columns or [])
        + details_value_columns
        + details_zscore_columns
        + details_metric_columns
    )
    outlier_cells = outlier_cells or {}
    trend_month_columns = trend_month_columns or []
    details_outlier_cells = details_outlier_cells or {}

    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_values = [cell.value for cell in ws[1]]
        material_col_idx = (
            header_values.index("Material") + 1 if "Material" in header_values else None
        )
        currency_col_indices = {
            header_values.index(col) + 1
            for col in currency_columns
            if col in header_values
        }
        percent_col_indices = {
            header_values.index(col) + 1
            for col in percent_columns
            if col in header_values
        }
        decimal_col_indices = {
            header_values.index(col) + 1
            for col in decimal_columns
            if col in header_values
        }

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

            if col_idx == material_col_idx:
                for cell in column_cells[1:]:  # Kopfzeile auslassen
                    cell.number_format = "@"
            elif col_idx in currency_col_indices:
                for cell in column_cells[1:]:
                    cell.number_format = CURRENCY_FORMAT
            elif col_idx in percent_col_indices:
                for cell in column_cells[1:]:
                    cell.number_format = PERCENT_FORMAT
            elif col_idx in decimal_col_indices:
                for cell in column_cells[1:]:
                    cell.number_format = COVERAGE_DECIMAL_FORMAT

        # Sondereffekt-Zellen farblich markieren (nur im Trend_MVER-Reiter,
        # nur fuer die rohen Monatsspalten - siehe Funktions-Docstring).
        if (
            sheet_name == TREND_SHEET_NAME
            and outlier_cells
            and material_col_idx is not None
        ):
            month_col_indices = {
                col: header_values.index(col) + 1
                for col in trend_month_columns
                if col in header_values
            }
            for row_idx in range(2, ws.max_row + 1):
                material_value = ws.cell(row=row_idx, column=material_col_idx).value
                affected_months = outlier_cells.get(material_value)
                if not affected_months:
                    continue
                for month_col in affected_months:
                    col_idx = month_col_indices.get(month_col)
                    if col_idx is not None:
                        ws.cell(row=row_idx, column=col_idx).fill = OUTLIER_CELL_FILL

        # Sondereffekt-Zellen farblich markieren (Reiter Trend_Details, ab
        # 25.06.2026) - dieselbe Logik wie oben fuer Trend_MVER, aber auf
        # den dortigen Wertspalten (details_value_columns), NICHT auf den
        # zugehoerigen Z-Score-Spalten (siehe Funktions-Docstring: der
        # Z-Score macht den Ausreisser bereits an seinem Betrag erkennbar,
        # eine zusaetzliche Markierung dort waere redundant).
        if (
            sheet_name == TREND_DETAILS_SHEET_NAME
            and details_outlier_cells
            and material_col_idx is not None
        ):
            details_month_col_indices = {
                col: header_values.index(col) + 1
                for col in details_value_columns
                if col in header_values
            }
            for row_idx in range(2, ws.max_row + 1):
                material_value = ws.cell(row=row_idx, column=material_col_idx).value
                affected_months = details_outlier_cells.get(material_value)
                if not affected_months:
                    continue
                for month_col in affected_months:
                    col_idx = details_month_col_indices.get(month_col)
                    if col_idx is not None:
                        ws.cell(row=row_idx, column=col_idx).fill = OUTLIER_CELL_FILL

        # Prioritaets-Spalte farblich markieren (Reiter Executive_Summary,
        # Phase 6, ab 29.06.2026) - eigene kleine Funktion statt weiterer
        # Sonderfaelle direkt in dieser Schleife (siehe Docstring von
        # _apply_priority_highlighting()).
        if sheet_name == EXECUTIVE_SUMMARY_SHEET_NAME:
            _apply_priority_highlighting(ws, header_values)

        ws.freeze_panes = "A2"

    wb.save(file_path)
